'''
Created on 2019/04/13

@author: matsumura
'''
import copy
import functools
import operator
import csv
import timeit
import openpyxl as excel
import cplex
from cplex.exceptions import CplexError
import function as fc

def experiment2(file_path, n, k_para):
    time_pcsg = []
    value_pcsg = []
    with open(file_path,'r') as f:
        reader=csv.reader(f)
        for row in reader:
            row.pop(-1)
            my_obj = [float(i) for i in row[:-n]]
            my_prob = [float(i) for i in row[-n:]]

            #Start solving the PCSG problems.
            start_pcsg  = timeit.default_timer()

            optimal_value, k_values = fc.solve_pcsg(n, k_para,
                                                    my_prob, my_obj)

            stop_pcsg = timeit.default_timer()
            t_pcsg = stop_pcsg - start_pcsg

            time_pcsg.append(t_pcsg)
            value_pcsg.append(optimal_value)

    return time_pcsg, value_pcsg

def create_result_file2(file_path, result_file_name, k_para_list):

    for i in range(len(file_list)):
        file_path = file_list[i]
        n = int(''.join(c for c in file_path if c.isdigit()))

        wb = excel.load_workbook(result_file_name)
        sheet_name = file_path.split('_')[1].split('.csv')[0]
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name, 0)
        ws = wb.active
        ws = wb[sheet_name]

        for i in range(len(k_para_list)):
            k_para = k_para_list[i]
            time_pcsg, value_pcsg = experiment2(file_path, n, k_para)

            #make a average runtime table
            ws.cell(column=1, row=1, value=sheet_name)
            ws.cell(column=k_para+1, row=2, value='k=%d' % k_para)
            ws.cell(column=1, row=n/2, value='n=%d' % n)
            ws.cell(column=k_para+1, row=n/2, value=sum(time_pcsg)/len(time_pcsg))
            ws.cell(column=2*n-4+k_para, row=1, value='n=%dk=%d' % (n, k_para))
            #plot runtime in a column
            for j in range(len(time_pcsg)):
                ws.cell(column=2*n-2+i, row=2+j, value=time_pcsg[j])

        wb.save(result_file_name)

    return
