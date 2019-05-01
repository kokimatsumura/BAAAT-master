'''
Created on 2019/04/16

@author: matsumura
'''
import copy
import functools
import operator
import csv
import timeit
import openpyxl as excel
import cplex
from cplex.exceptions import CplexError
import function as fc

def experiment3(file_path, n, k_para, p_tilde):
    time_baaat = []
    value_baaat = []
    with open(file_path,'r') as f:
        reader=csv.reader(f)
        for row in reader:
            row.pop(-1)
            my_obj = [float(i) for i in row[:-n]]
            my_prob = [float(i) for i in row[-n:]]

            #Start solving the BAAAT problems.
            start_baaat = timeit.default_timer()

            optimal_value_baaat, reduced_list\
                    = fc.solve_pcsg_by_baaat(n, k_para, p_tilde,
                                                my_prob, my_obj)

            stop_baaat = timeit.default_timer()
            t_baaat = stop_baaat - start_baaat
            time_baaat.append(t_baaat)
            value_baaat.append(optimal_value_baaat)

    return time_baaat, value_baaat

def create_result_file3(file_path, result_file_name, p_tilde_list, k_para):

    for i in range(len(file_list)):
        file_path = file_list[i]
        n = int(''.join(c for c in file_path if c.isdigit()))

        wb = excel.load_workbook(result_file_name)
        sheet_name = file_path.split('_')[1].split('.csv')[0]
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name, 0)

        ws = wb.active
        ws = wb[sheet_name]

        for i in range(len(p_tilde_list)):
            p_tilde = p_tilde_list[i]
            time_baaat, value_baaat = experiment3(file_path, n,
                                                k_para, p_tilde)
            #make a average runtime table
            ws.cell(column=1, row=1, value=sheet_name)
            ws.cell(column=i+2, row=2, value='p_tilde=%01.1f' % p_tilde)
            ws.cell(column=1, row=n/2, value='n=%d' % n)
            ws.cell(column=i+2, row=n/2, value=sum(time_baaat)/len(time_baaat))
            ws.cell(column=2*n-4+i, row=1, value='n=%dp_tilde=%01.1f' % (n, p_tilde))
            #plot runtime in a column
            for j in range(len(time_baaat)):
                ws.cell(column=2*n-4+i, row=2+j, value=time_baaat[j])

            #make a average value table
            ws.cell(column=i+2, row=2+10, value='p_tilde=%01.1f' % p_tilde)
            ws.cell(column=1, row=n/2+10, value='n=%d' % n)
            ws.cell(column=i+2, row=n/2+10, value=sum(value_baaat)/len(value_baaat))

            #plot values in a column
            ws.cell(column=2*n-4+i, row=3+len(time_baaat), value='n=%dp_tilde=%01.1f' % (n, p_tilde))

            for j in range(len(value_baaat)):
                ws.cell(column=2*n-4+i, row=4+len(time_baaat)+j, value=value_baaat[j])

        wb.save(result_file_name)

    return
