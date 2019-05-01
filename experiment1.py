'''
Created on 2019/04/09

@author: matsumura
'''
import copy
import functools
import operator
import csv
import timeit
import openpyxl as excel
import cplex
from cplex.exceptions import CplexError
import function as fc

def experiment1(file_path, n, k_para, p_tilde):
    time_baaat = []
    value_baaat = []
    time_pcsg = []
    value_pcsg = []
    ub_list = []
    with open(file_path,'r') as f:
        reader=csv.reader(f)
        for row in reader:
            row.pop(-1)
            my_obj = [float(i) for i in row[:-n]]
            my_prob = [float(i) for i in row[-n:]]

            #Start solving the BAAAT problems.
            start_baaat = timeit.default_timer()

            optimal_value_baaat, reduced_list = fc.solve_pcsg_by_baaat(n, k_para, p_tilde,
                                                                       my_prob, my_obj)

            stop_baaat = timeit.default_timer()
            t_baaat = stop_baaat - start_baaat
            time_baaat.append(t_baaat)
            value_baaat.append(optimal_value_baaat)
            #Start solving the PCSG problems.
            start_pcsg  = timeit.default_timer()

            optimal_value, k_values = fc.solve_pcsg(n, k_para,
                                                    my_prob, my_obj)

            stop_pcsg = timeit.default_timer()
            t_pcsg = stop_pcsg - start_pcsg

            #Calcurate the upper bounds
            error_bound = fc.calc_error_bound(reduced_list, k_values)

            time_pcsg.append(t_pcsg)
            value_pcsg.append(optimal_value)
            ub_list.append(error_bound)

    return time_baaat, value_baaat, time_pcsg, value_pcsg, ub_list

def create_result_file1(file_list, result_file_name, k_para, p_tilde):

    for i in range(len(file_list)):
        file_path = file_list[i]
        n = int(file_path.split('=')[1].split('_')[0])

        time_baaat, value_baaat, time_pcsg, value_pcsg, ub_list\
                                = experiment(file_path, n, k_para, p_tilde)

        wb = excel.load_workbook(result_file_name)
        sheet_name = file_path.split('_')[1].split('.csv')[0]
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name, 0)

        ws = wb.active
        ws = wb[sheet_name]
        ws.cell(column=7, row=1, value=sheet_name)
        ws.cell(column=1, row=1, value='BAAAT-time')
        ws.cell(column=2, row=1, value='BAAAT-value')
        ws.cell(column=3, row=1, value='PCSG-time')
        ws.cell(column=4, row=1, value='PCSG-value')
        ws.cell(column=5, row=1, value='UB')
        ws.cell(column=1, row=(n/2 - 3) * (len(time_baaat)+1) + 2,
                    value='n=%dk=%dp_tilde=%01.1f' % (n, k_para, p_tilde))

        for j in range(len(time_baaat)):
            count = (n/2 - 3) * (len(time_baaat)+1) + 3 + j
            ws.cell(column=1, row=count, value=time_baaat[j])
            ws.cell(column=2, row=count, value=value_baaat[j])
            ws.cell(column=3, row=count, value=time_pcsg[j])
            ws.cell(column=4, row=count, value=value_pcsg[j])
            ws.cell(column=5, row=count, value=ub_list[j])

        #make a value table
        ws.cell(column=7, row=2, value='number of agents')
        ws.cell(column=n/2 + 5, row=2, value='n=%d' %n)
        ws.cell(column=7, row=3, value='complete algorithm')
        ws.cell(column=7, row=4, value='BAAAT')
        ws.cell(column=7, row=5, value='ub')
        ws.cell(column=n/2 + 5, row=3, value=sum(value_pcsg)/len(value_pcsg))
        ws.cell(column=n/2 + 5, row=4, value=sum(value_baaat)/len(value_baaat))
        ws.cell(column=n/2 + 5, row=5, value=sum(ub_list)/len(ub_list))

        #make a runtime table
        ws.cell(column=7, row=8, value='runtime')
        ws.cell(column=n/2 + 5, row=8, value='n=%d' %n)
        ws.cell(column=7, row=9, value='complete algorithm')
        ws.cell(column=7, row=10, value='BAAAT')
        ws.cell(column=n/2 + 5, row=9, value=sum(time_pcsg)/len(time_pcsg))
        ws.cell(column=n/2 + 5, row=10, value=sum(time_baaat)/len(time_baaat))

        #make a quality table
        avg_pcsg = sum(value_pcsg)/len(value_pcsg)
        avg_baaat = sum(value_baaat)/len(value_baaat)
        avg_ub = sum(ub_list)/len(ub_list)

        ws.cell(column=7, row=13, value='number of agents')
        ws.cell(column=n/2 + 5, row=13, value='n=%d' %n)
        ws.cell(column=7, row=14, value='BAAAT')
        ws.cell(column=7, row=15, value='ub')
        ws.cell(column=n/2 + 5, row=14, value=avg_baaat / avg_pcsg)
        ws.cell(column=n/2 + 5, row=15, value=(avg_baaat + avg_ub) / avg_pcsg)

        wb.save(result_file_name)

    return
